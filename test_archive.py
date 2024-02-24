import zipfile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook
from settings import archive
from io import TextIOWrapper


def test_csv():
    """Проверка содержимого csv-файла в архиве"""
    with zipfile.ZipFile(archive, 'r', zipfile.ZIP_DEFLATED) as zp:
        with zp.open('file.csv', 'r', ) as csv_f:
            reader = list(csv.reader(TextIOWrapper(csv_f, 'utf-8-sig'), delimiter=';'))
            second_row = reader[1]

    assert second_row[1] == 'Dulce'  # проверка значения элемента в первом столбце второй строки
    assert second_row[2] == 'Abril'  # проверка значения элемента во втором столбце второй строки


def test_pdf():
    """Проверка содержимого pdf-файла в архиве"""
    with zipfile.ZipFile(archive, 'r', zipfile.ZIP_DEFLATED) as zp:
        with zp.open('file.pdf', 'r', ) as pdf_f:
            reader = PdfReader(pdf_f)
            author = reader.metadata.author
            sheets_count = len(reader.pages)

    assert author == "Ёшкин Кот"  # проверка автора
    assert sheets_count == 69  # проверка количества страниц


def test_xlsx():
    """Проверка содержимого xlsx-файла в архиве"""
    with zipfile.ZipFile(archive, 'r', zipfile.ZIP_DEFLATED) as zp:
        with zp.open('file.xlsx', 'r', ) as xlsx_f:
            workbook = load_workbook(xlsx_f)
            worksheet = workbook.active
            first_name = worksheet.cell(row=3, column=2).value
            last_name = worksheet.cell(row=3, column=3).value
            rows_count = worksheet.max_row

    assert first_name == 'Mara'
    assert last_name == 'Hashimoto'
    assert rows_count == 51
